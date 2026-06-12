"""Export the Customer Risk Register to an Excel workbook.

Read-only: pulls the receivables ``customers`` table from Supabase (the same
table the live dashboard's Risk Register reads) and writes an **.xlsx** with two
sheets:

  • "Risk Register" — one row per ``(customer, company, location)`` with Company
    and Location as their own columns, the full aging breakdown (0-30 … 180+) as
    separate columns alongside the total Overdue, and a derived **Blocked** flag.
  • "By Sale Type" — one row per ``(customer, company, location, sale type)``,
    with the type-sliceable figures (Sales/Receipts/Credit Notes/Outstanding/
    Overdue/aging) taken from the ``*_by_type`` JSON maps.

No Supabase writes; nothing in Tally is touched. The ``customers`` table already
stores pre-computed ``aging_buckets`` / ``*_by_type`` JSON per row, so this script
only reads, flattens, and formats — it does not recompute aging.

Notes on two fields that are NOT stored directly:
  • Blocked is derived: ``credit_limit == 1`` (the INK sentinel the dashboard uses).
  • A customers row has no single sale type — it aggregates all types, with per-type
    breakdowns in the ``*_by_type`` maps. Opening / Debit Notes / Journal have no
    per-type split, so they live on the "Risk Register" sheet only.

Required env vars (loaded from project .env):
    SUPABASE_URL           e.g. https://<ref>.supabase.co (no /rest/v1 suffix)
    SUPABASE_SERVICE_KEY   service_role / secret key (bypasses RLS)

Usage:
    python export_risk_register.py
    python export_risk_register.py --fiscal-year default --out-dir ../../MISC
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from supabase import create_client, Client

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# tools/sync_dashboard/ -> project root is two levels up.
PROJECT_ROOT = Path(__file__).resolve().parents[2]

# PostgREST caps a single response at 1000 rows; page through with .range().
PAGE_SIZE = 1000

# Internal aging-bucket keys (in aging_buckets JSON) -> human column header.
AGING_BUCKETS: list[tuple[str, str]] = [
    ("0_30", "0-30"),
    ("31_60", "31-60"),
    ("61_90", "61-90"),
    ("91_120", "91-120"),
    ("121_180", "121-180"),
    ("180_plus", "180+"),
]

# Sale-type keys (in the *_by_type JSON maps) -> display label. Order = row order
# on the "By Sale Type" sheet; unknown keys are appended (sorted) with a derived label.
SALE_TYPE_ORDER: list[str] = ["machine", "ink", "spare_parts", "head", "other"]
SALE_TYPE_LABEL: dict[str, str] = {
    "machine": "Machine",
    "ink": "Ink",
    "spare_parts": "Spare Parts",
    "head": "Head",
    "other": "Other",
}

# (header, source-row key, is_amount) — shared base + tail column groups.
BASE_COLUMNS: list[tuple[str, str, bool]] = [
    ("Customer", "name", False),
    ("Company", "company", False),
    ("Location", "location", False),
    ("Sales Person", "sales_person", False),
    ("Opening", "opening_balance", True),
    ("Sales", "sales", True),
    ("Receipts", "receipts", True),
    ("Credit Notes", "credit_notes", True),
    ("Debit Notes", "debit_notes", True),
    ("Journal (Net)", "journal_adjustments", True),
    ("Outstanding", "outstanding", True),
    ("Overdue", "overdue", True),
]

# Customer-level descriptors repeated on every row (they describe the ledger, not a slice).
TAIL_COLUMNS: list[tuple[str, str, bool]] = [
    ("Max OD Days", "max_overdue_days", False),
    ("Credit Period", "credit_period", False),
    ("Credit Limit", "credit_limit", True),
    ("Util %", "utilization", False),
    ("Risk", "risk", False),
]

# Excel number format for rupee amounts: thousands separators, no decimals.
AMOUNT_FMT = "#,##0"


def _to_number(v: Any) -> float | int:
    """Coerce a Supabase numeric/string to a number; blanks/garbage -> 0."""
    if v is None or v == "":
        return 0
    try:
        f = float(v)
    except (TypeError, ValueError):
        return 0
    return int(f) if f == int(f) else f


def _cell(row: dict[str, Any], key: str, is_amount: bool) -> Any:
    """Render a customer-row value for a cell: numbers for amounts, '' for blanks."""
    if is_amount:
        return _to_number(row.get(key))
    v = row.get(key)
    return "" if v is None else v


def _json_map(row: dict[str, Any], col: str) -> dict[str, Any]:
    """Return a JSON-object column as a dict (parse strings, default {})."""
    raw = row.get(col)
    if isinstance(raw, str):
        try:
            raw = json.loads(raw)
        except json.JSONDecodeError:
            raw = {}
    return raw if isinstance(raw, dict) else {}


def _aging(buckets: dict[str, Any]) -> dict[str, float | int]:
    """Return the six bucket values from an aging_buckets dict, missing -> 0."""
    return {key: _to_number(buckets.get(key)) for key, _ in AGING_BUCKETS}


def _blocked_str(row: dict[str, Any]) -> str:
    """Blocked flag = the INK credit-limit sentinel (credit_limit == 1)."""
    return "Yes" if _to_number(row.get("credit_limit")) == 1 else "No"


def fetch_customers(client: Client, fiscal_year: str) -> list[dict[str, Any]]:
    """Page through the customers table for one fiscal-year snapshot."""
    rows: list[dict[str, Any]] = []
    offset = 0
    while True:
        resp = (
            client.table("customers")
            .select("*")
            .eq("fiscal_year", fiscal_year)
            .range(offset, offset + PAGE_SIZE - 1)
            .execute()
        )
        batch = resp.data or []
        rows.extend(batch)
        if len(batch) < PAGE_SIZE:
            break
        offset += PAGE_SIZE
    return rows


# ── sheet writer ──────────────────────────────────────────────────────────────

def write_sheet(ws, headers: list[str], amount_flags: list[bool], data_rows: list[list[Any]]) -> None:
    """Write a header + rows to a worksheet with bold/frozen header, amount number
    formats, and auto column widths."""
    ws.append(headers)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for values in data_rows:
        ws.append(values)

    widths = [len(h) for h in headers]
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            is_amount = amount_flags[c - 1]
            if is_amount and isinstance(cell.value, (int, float)):
                cell.number_format = AMOUNT_FMT
                text = f"{cell.value:,.0f}"
            else:
                text = str(cell.value if cell.value is not None else "")
            if len(text) > widths[c - 1]:
                widths[c - 1] = len(text)

    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = min(max(w + 2, 8), 48)

    ws.freeze_panes = "A2"


# ── sheet builders ────────────────────────────────────────────────────────────

def build_register_rows(rows: list[dict[str, Any]]) -> tuple[list[str], list[bool], list[list[Any]]]:
    """Sheet 1 — per (customer, company, location) aggregate + Blocked."""
    headers = (
        [h for h, _, _ in BASE_COLUMNS]
        + [label for _, label in AGING_BUCKETS]
        + ["Aging Total"]
        + [h for h, _, _ in TAIL_COLUMNS]
        + ["Blocked"]
    )
    amount_flags = (
        [amt for _, _, amt in BASE_COLUMNS]
        + [True] * len(AGING_BUCKETS)
        + [True]                                  # Aging Total
        + [amt for _, _, amt in TAIL_COLUMNS]
        + [False]                                 # Blocked
    )

    data: list[list[Any]] = []
    for row in rows:
        buckets = _aging(_json_map(row, "aging_buckets"))
        values: list[Any] = [_cell(row, key, amt) for _, key, amt in BASE_COLUMNS]
        values.extend(buckets[key] for key, _ in AGING_BUCKETS)
        values.append(sum(buckets.values()))      # Aging Total
        values.extend(_cell(row, key, amt) for _, key, amt in TAIL_COLUMNS)
        values.append(_blocked_str(row))
        data.append(values)
    return headers, amount_flags, data


def _all_sale_types(rows: list[dict[str, Any]]) -> list[str]:
    """Global sale-type set (union of sales_by_type keys), matching the dashboard's
    `allSaleTypes` (useAppData.ts:491). Ordered by SALE_TYPE_ORDER, extras appended."""
    seen: set[str] = set()
    for row in rows:
        seen |= set(_json_map(row, "sales_by_type").keys())
    ordered = [t for t in SALE_TYPE_ORDER if t in seen]
    ordered += sorted(t for t in seen if t not in SALE_TYPE_ORDER)
    return ordered


def build_by_sale_type_rows(rows: list[dict[str, Any]]) -> tuple[list[str], list[bool], list[list[Any]]]:
    """Sheet 2 — one row per (customer, company, location, sale type).

    Replicates the dashboard's per-type *projection* (useAppData.ts:541-573): each
    type's figure = its typed portion + a share of the untyped residual (opening
    balance, advances, unlinked credit notes, cheque returns, Tally delta), where the
    share is that type's fraction of the customer's sales (residual lands in "other"
    when the customer has no sales). This makes the per-type rows reconcile to the
    Sheet-1 aggregate and matches what the dashboard shows when filtering by Sale Type.
    (Aging Total may not equal Overdue per row — same minor source-data quirk the
    dashboard carries; the reconciliation guarantee is the per-type SUM = the total.)
    Max OD Days / Risk / Util %% are customer-level (not type-specific) and repeated.
    """
    headers = (
        ["Customer", "Company", "Location", "Sales Person", "Sale Type"]
        + ["Sales", "Receipts", "Credit Notes", "Outstanding", "Overdue"]
        + [label for _, label in AGING_BUCKETS]
        + ["Aging Total"]
        + [h for h, _, _ in TAIL_COLUMNS]
        + ["Blocked"]
    )
    amount_flags = (
        [False] * 5                                # identity + Sale Type
        + [True] * 5                               # Sales..Overdue
        + [True] * len(AGING_BUCKETS)
        + [True]                                   # Aging Total
        + [amt for _, _, amt in TAIL_COLUMNS]
        + [False]                                  # Blocked
    )

    all_types = _all_sale_types(rows)
    EPS = 0.5  # drop a type row whose every projected figure rounds to 0

    data: list[list[Any]] = []
    for row in rows:
        sales_bt = _json_map(row, "sales_by_type")
        receipts_bt = _json_map(row, "receipts_by_type")
        cn_bt = _json_map(row, "credit_notes_by_type")
        out_bt = _json_map(row, "outstanding_by_type")
        over_bt = _json_map(row, "overdue_by_type")
        aging_bt = _json_map(row, "aging_buckets_by_type")
        agg_buckets = _aging(_json_map(row, "aging_buckets"))

        # Customer totals + the typed sums whose gap to the total is the residual.
        def typed_sum(m: dict[str, Any]) -> float:
            return sum(_to_number(m.get(tt)) for tt in all_types)

        res_out = _to_number(row.get("outstanding")) - typed_sum(out_bt)
        res_over = _to_number(row.get("overdue")) - typed_sum(over_bt)
        res_rec = _to_number(row.get("receipts")) - typed_sum(receipts_bt)
        res_cn = _to_number(row.get("credit_notes")) - typed_sum(cn_bt)
        res_bucket = {
            key: agg_buckets[key] - sum(
                _to_number((aging_bt.get(tt) or {}).get(key)) for tt in all_types
            )
            for key, _ in AGING_BUCKETS
        }

        sales_total = typed_sum(sales_bt)
        has_sales = sales_total > 1e-9

        blocked = _blocked_str(row)
        tail = [_cell(row, key, amt) for _, key, amt in TAIL_COLUMNS]

        for t in all_types:
            share = (_to_number(sales_bt.get(t)) / sales_total) if has_sales else (1.0 if t == "other" else 0.0)

            sales = _to_number(sales_bt.get(t))
            receipts = _to_number(receipts_bt.get(t)) + res_rec * share
            credit_notes = _to_number(cn_bt.get(t)) + res_cn * share
            outstanding = _to_number(out_bt.get(t)) + res_out * share
            overdue = _to_number(over_bt.get(t)) + res_over * share
            type_buckets = aging_bt.get(t) or {}
            buckets = {
                key: _to_number(type_buckets.get(key)) + res_bucket[key] * share
                for key, _ in AGING_BUCKETS
            }
            aging_total = sum(buckets.values())

            if all(abs(v) < EPS for v in
                   (sales, receipts, credit_notes, outstanding, overdue, aging_total)):
                continue

            label = SALE_TYPE_LABEL.get(t, t.replace("_", " ").title())
            values: list[Any] = [
                row.get("name") or "", row.get("company") or "", row.get("location") or "",
                row.get("sales_person") or "", label,
                sales, receipts, credit_notes, outstanding, overdue,
            ]
            values.extend(buckets[key] for key, _ in AGING_BUCKETS)
            values.append(aging_total)
            values.extend(tail)
            values.append(blocked)
            data.append(values)
    return headers, amount_flags, data


def build_workbook(rows: list[dict[str, Any]]) -> tuple[Workbook, int]:
    """Build the two-sheet workbook; return (wb, sale_type_row_count)."""
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Risk Register"
    h1, f1, d1 = build_register_rows(rows)
    write_sheet(ws1, h1, f1, d1)

    ws2 = wb.create_sheet("By Sale Type")
    h2, f2, d2 = build_by_sale_type_rows(rows)
    write_sheet(ws2, h2, f2, d2)

    return wb, len(d2)


def main() -> int:
    parser = argparse.ArgumentParser(description="Export the Customer Risk Register to Excel.")
    parser.add_argument("--fiscal-year", default="default",
                        help="customers.fiscal_year snapshot to export (default / fy2526 / fy2627)")
    parser.add_argument("--out-dir", default=str(PROJECT_ROOT / "MISC"),
                        help="Directory to write the .xlsx into (default: <project>/MISC)")
    args = parser.parse_args()

    load_dotenv()
    url = os.environ["SUPABASE_URL"].strip().rstrip("/")
    key = os.environ["SUPABASE_SERVICE_KEY"].strip()
    if url.endswith("/rest/v1"):
        url = url[: -len("/rest/v1")]

    client = create_client(url, key)
    rows = fetch_customers(client, args.fiscal_year)

    wb, type_rows = build_workbook(rows)

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%d-%m-%Y")
    out_path = out_dir / f"risk-register-{stamp}.xlsx"
    wb.save(out_path)

    print(json.dumps({
        "rows": len(rows),
        "type_rows": type_rows,
        "file": str(out_path),
        "fiscal_year": args.fiscal_year,
    }))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
