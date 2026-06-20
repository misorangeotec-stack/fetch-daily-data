"""One-off importer: parse the manual "OTHER PAYMENTS" Excel tab into a flat
Google Sheet that the Orange Receivables Hub pipeline can read.

The source workbook (MISC/SALES PERSON - DETAILS.xlsx, tab "OTHER PAYMENTS") is a
GROUPED human sheet, not a flat table:

    A=invoice date  B=invoice/bill ref  C=customer name (header rows only)
    D=base amt      E=GST               F=total  (= the payment amount)
    G=payment date  H=short payment ref I=salesperson

Row kinds (classified top-to-bottom, carrying the current customer forward):
  * header row     — C has a customer name, F empty            → set current customer
  * detail row     — C empty, F present, A or B present        → a payment line
  * subtotal row   — A/B/C empty, D/E/F present                → skip
  * grand-total    — A/B/C empty                               → skip
  * on-account tail— customer name in col A, amount in col D    → ON_ACCOUNT line
                     (ETHOS / SINO STAR / VERONICA DIGITAL)
  * TWINE-style    — detail row whose ref is a bare number     → ON_ACCOUNT

Allocation rule: a ref that looks like a real invoice (contains "/" or a letter)
→ AGAINST_INVOICE; anything else (blank, bare number) → ON_ACCOUNT.

Output sheet columns (flat):
    Company | Location | Customer Name | Ref Invoice No | Payment Amount |
    Payment Date (dd-mm-yyyy) | Payment Ref | Allocation Type | Salesperson

Company/Location are left blank — these rows are all Otec Surat and the pipeline
resolves them by name-only (see process_data.py name-only resolver).

Usage:
    python import_other_payments.py                      # parse + create a NEW sheet
    python import_other_payments.py --dry-run            # parse + print, write nothing
    python import_other_payments.py --sheet-url <url>    # write into an existing sheet/tab
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from dotenv import load_dotenv
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
SHEET_ID_RE = re.compile(r"/spreadsheets/d/([a-zA-Z0-9-_]+)")

DEFAULT_XLSX = Path(__file__).resolve().parent / "SALES PERSON - DETAILS.xlsx"
SRC_TAB = "OTHER PAYMENTS"

HEADER = [
    "Company", "Location", "Customer Name", "Ref Invoice No", "Payment Amount",
    "Payment Date", "Payment Ref", "Allocation Type", "Salesperson", "Remarks",
]

# Header-cell notes that are NOT a salesperson (col I sometimes carries a remark).
NON_SALESPERSON_NOTES = {"PAYMENT RECEIVED"}

# Loose names in the manual sheet → the exact customer-master ledger name, so the
# pipeline's name resolver can match them. Keyed by UPPER+TRIM of the sheet name.
# "sino star" → the base trading ledger "SINO STAR ENTERPRISE" (consistent with the
# other on-account entries ETHOS / VERONICA DIGITAL using their base, non-machine ledger).
CUSTOMER_ALIASES = {
    "SINO STAR": "SINO STAR ENTERPRISE",
}


def _canon_customer(name: str) -> str:
    return CUSTOMER_ALIASES.get(name.strip().upper(), name.strip())


# Explicit (Company, Location) for customers the pipeline can't resolve automatically.
# Many customers trade under several company books (O-tec / Enterprise / Colorix ×
# Surat / Noida) — a separate Tally ledger (GUID) each. The pipeline resolves an
# AGAINST_INVOICE row by the bill's owning ledger, but that fails when the payment is
# on account (no invoice) OR the referenced invoice predates the current sales feed /
# uses a reused voucher number. For those, pin the ledger here. All listed below resolve
# to O-tec / Surat — the entity that actually carries each customer's receivable & sales
# (verified) and matches the "Otec Surat" classification. Edit to retarget a payment.
# Keyed by UPPER+TRIM of the canonical customer name.
CUSTOMER_LEDGER = {
    # on-account (no invoice ref)
    "ETHOS":                  ("O-tec", "Surat"),
    "SINO STAR ENTERPRISE":   ("Enterprise", "Surat"),  # settle the 25L against the Enterprise outstanding
    "VERONICA DIGITAL":       ("O-tec", "Surat"),
    # against-invoice but ref not resolvable (old/opening invoice or reused voucher #)
    "ADIYA DESIGNER":         ("O-tec", "Surat"),
    "PEACOCK DIGITAL PRINTS": ("O-tec", "Surat"),
    "SHIVRAM PROCESSORS":     ("O-tec", "Surat"),
}
# A ref that looks like a real Tally invoice: has a slash or a letter.
INVOICE_REF_RE = re.compile(r"[A-Za-z/]")

# FY 25-26 start: invoices dated before this are prior-year (rolled into the opening
# balance), so a payment against them can't target a current bill -> booked On Account.
FY2526_CUTOFF = date(2025, 4, 1)


def _norm(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _is_blank_amt(val: Any) -> bool:
    """Treat None / "" / 0 as empty for header-vs-detail classification."""
    if val is None or val == "":
        return True
    try:
        return float(val) == 0.0
    except (TypeError, ValueError):
        return False


def _fmt_date(val: Any) -> str:
    if isinstance(val, datetime):
        return val.strftime("%d-%m-%Y")
    if isinstance(val, date):
        return val.strftime("%d-%m-%Y")
    s = _norm(val)
    if not s:
        return ""
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).strftime("%d-%m-%Y")
        except ValueError:
            continue
    return ""


def _fmt_amount(val: Any) -> str:
    try:
        return str(round(float(val), 2))
    except (TypeError, ValueError):
        return ""


def _alloc_for(ref: str) -> str:
    """AGAINST_INVOICE if ref looks like a real invoice, else ON_ACCOUNT."""
    if ref and INVOICE_REF_RE.search(ref):
        return "AGAINST_INVOICE"
    return "ON_ACCOUNT"


def parse_other_payments(xlsx_path: Path) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SRC_TAB not in wb.sheetnames:
        raise SystemExit(f"ERROR: tab '{SRC_TAB}' not found in {xlsx_path}")
    ws = wb[SRC_TAB]

    rows: list[dict[str, str]] = []
    current_customer = ""
    current_salesperson = ""

    for r in ws.iter_rows(min_row=1, values_only=True):
        # pad to 9 cols (A..I)
        cells = list(r) + [None] * (9 - len(r))
        a, b, c, d, e, f, g, h, i = cells[:9]

        c_name = _norm(c)
        a_str = _norm(a) if not isinstance(a, (datetime, date)) else ""
        amt_f_blank = _is_blank_amt(f)

        # ── on-account tail: a customer name sits in col A (a string, not a date).
        # These rows carry the amount in col D (col C may hold a note like "BANK
        # PAYMENT"), so check this BEFORE the header rule. ──
        if a_str:
            amt = f if not amt_f_blank else d
            if not _is_blank_amt(amt):
                rows.append({
                    "Company": "", "Location": "",
                    "Customer Name": _canon_customer(a_str),
                    "Ref Invoice No": "",
                    "Payment Amount": _fmt_amount(amt),
                    "Payment Date": _fmt_date(g),
                    "Payment Ref": _norm(h),
                    "Allocation Type": "ON_ACCOUNT",
                    "Salesperson": _norm(i),
                    "Remarks": "",
                    "_invoice_date": None,
                })
            continue

        # ── header row: customer name in C, no total → set context ──
        if c_name and amt_f_blank and _is_blank_amt(b):
            current_customer = c_name
            sp = _norm(i).upper()
            current_salesperson = "" if sp in NON_SALESPERSON_NOTES else _norm(i)
            continue

        # ── subtotal / grand-total / blank rows: skip ──
        ref = _norm(b)
        has_invoice_date = isinstance(a, (datetime, date))
        if not c_name and not has_invoice_date and not ref:
            continue  # subtotal (amounts only) or empty

        # ── detail row: a payment line for the current customer ──
        if amt_f_blank:
            continue  # no payment amount → not a real line
        if not current_customer:
            continue  # safety: no customer context

        salesperson = _norm(i) or current_salesperson
        rows.append({
            "Company": "", "Location": "",
            "Customer Name": _canon_customer(current_customer),
            "Ref Invoice No": ref,
            "Payment Amount": _fmt_amount(f),
            "Payment Date": _fmt_date(g),
            "Payment Ref": _norm(h),
            "Allocation Type": _alloc_for(ref),
            "Salesperson": salesperson,
            "Remarks": "",
            "_invoice_date": a if isinstance(a, (datetime, date)) else None,
        })

    # On-account refs are placeholders (e.g. TWINE "1") — blank them out.
    for row in rows:
        if row["Allocation Type"] == "ON_ACCOUNT":
            row["Ref Invoice No"] = ""
    return rows


# ── Ledger resolution: fill Company/Location, mark old invoices on-account ──────

def _load_values(svc: Any, url_env: str, tab_env: str) -> tuple[list[str], list[list[str]]]:
    url = os.environ.get(url_env, "").strip()
    if not url:
        return [], []
    tab = os.environ.get(tab_env, "Sheet1")
    resp = svc.spreadsheets().values().get(
        spreadsheetId=extract_sheet_id(url), range=f"'{tab}'", majorDimension="ROWS",
    ).execute()
    vals = resp.get("values", [])
    if not vals:
        return [], []
    return [h.strip() for h in vals[0]], vals[1:]


def build_resolution_maps(svc: Any) -> dict:
    """voucher/bill-ref -> owning ledgers, and the customer master -> ledgers."""
    voucher_map: dict[str, set] = {}   # VOUCHER -> {(name_u, company, location)}
    billref_map: dict[str, set] = {}   # BILL REF -> {(name_u, company, location)}

    def _idx(header, *names):
        low = [h.lower() for h in header]
        for n in names:
            if n in low:
                return low.index(n)
        return None

    # Sales: Voucher No. -> ledger
    sh, srows = _load_values(svc, "SALES_SHEET_URL", "SALES_SHEET_TAB")
    if sh:
        vi, ni, ci, li = (_idx(sh, "voucher no."), _idx(sh, "particulars"),
                          _idx(sh, "company"), _idx(sh, "location"))
        for r in srows:
            def g(idx): return r[idx].strip() if idx is not None and idx < len(r) else ""
            v = g(vi).upper()
            if v:
                voucher_map.setdefault(v, set()).add((g(ni).upper(), g(ci), g(li)))

    # Sales details (register): Bill Ref Name + Voucher No. -> ledger
    dh, drows = _load_values(svc, "SALES_DETAILS_SHEET_URL", "SALES_DETAILS_SHEET_TAB")
    if dh:
        bi, vi, ni, ci, li = (_idx(dh, "bill ref name"), _idx(dh, "voucher no."),
                              _idx(dh, "particulars"), _idx(dh, "company"), _idx(dh, "location"))
        for r in drows:
            def g(idx): return r[idx].strip() if idx is not None and idx < len(r) else ""
            led = (g(ni).upper(), g(ci), g(li))
            if g(bi):
                billref_map.setdefault(g(bi).upper(), set()).add(led)
            if g(vi):
                voucher_map.setdefault(g(vi).upper(), set()).add(led)

    # Customer master: $Name -> {(company, location)}
    master_map: dict[str, set] = {}
    mh, mrows = _load_values(svc, "CREDIT_LIMIT_SHEET_URL", "CREDIT_LIMIT_SHEET_TAB")
    if mh:
        ni = _idx(mh, "$name") or next((j for j, h in enumerate(mh) if "name" in h.lower()), None)
        ci, li = _idx(mh, "company"), _idx(mh, "location")
        for r in mrows:
            def g(idx): return r[idx].strip() if idx is not None and idx < len(r) else ""
            if ni is not None and g(ni):
                master_map.setdefault(g(ni).upper(), set()).add((g(ci), g(li)))

    return {"voucher": voucher_map, "billref": billref_map, "master": master_map}


def _lookup_invoice(ref: str, name_u: str, voucher_map: dict, billref_map: dict):
    """The (company, location) that owns this invoice ref FOR THIS CUSTOMER, else None.
    Requires the invoice's ledger name to match the payment's customer — a voucher number
    reused by another customer (or a stray ref) is NOT this customer's invoice."""
    r = ref.strip().upper()
    for mp in (voucher_map, billref_map):
        cands = mp.get(r)
        if not cands:
            continue
        named = {(co, lo) for (nm, co, lo) in cands if nm == name_u}
        if len(named) == 1:
            return next(iter(named))
    return None


def resolve_ledgers(rows: list[dict[str, str]], maps: dict) -> None:
    """Fill Company/Location for every row. An invoice that's found stays AGAINST_INVOICE
    on its owning ledger; a blank-ref or old/not-found invoice becomes ON_ACCOUNT with a
    remark, pinned to the customer's ledger (explicit override -> the company of the
    customer's matched invoices -> a unique master ledger)."""
    from collections import Counter
    voucher_map, billref_map, master_map = maps["voucher"], maps["billref"], maps["master"]

    # Pass 1: resolve found invoices; learn each customer's "active" company.
    matched: dict[str, Counter] = {}
    for row in rows:
        ref, name_u = row["Ref Invoice No"].strip(), row["Customer Name"].strip().upper()
        if ref and INVOICE_REF_RE.search(ref):
            colo = _lookup_invoice(ref, name_u, voucher_map, billref_map)
            if colo:
                row["_resolved"] = colo
                matched.setdefault(name_u, Counter())[colo] += 1
    primary = {n: c.most_common(1)[0][0] for n, c in matched.items()}

    # Pass 2: finalize.
    for row in rows:
        ref, name_u = row["Ref Invoice No"].strip(), row["Customer Name"].strip().upper()
        if row.pop("_resolved", None):
            row["Company"], row["Location"] = _lookup_invoice(ref, name_u, voucher_map, billref_map)
            row["Allocation Type"] = "AGAINST_INVOICE"
            row["Remarks"] = ""
            continue
        had_ref = bool(ref and INVOICE_REF_RE.search(ref))
        uniq = master_map.get(name_u)
        co, lo = (CUSTOMER_LEDGER.get(name_u)
                  or primary.get(name_u)
                  or (next(iter(uniq)) if uniq and len(uniq) == 1 else None)
                  or ("", ""))
        row["Company"], row["Location"] = co, lo
        row["Allocation Type"] = "ON_ACCOUNT"
        inv_date = row.get("_invoice_date")
        inv_d = inv_date.date() if isinstance(inv_date, datetime) else inv_date  # date | None
        if had_ref:
            r_u = ref.strip().upper()
            ref_exists = (r_u in voucher_map) or (r_u in billref_map)
            if inv_d is not None and inv_d < FY2526_CUTOFF:
                row["Remarks"] = (f"On Account - invoice {ref} dated {inv_d.strftime('%d-%m-%Y')} "
                                  f"is before 01-04-2025 (prior year / opening balance)")
            elif ref_exists:
                row["Remarks"] = f"On Account - invoice {ref} belongs to another ledger"
            else:
                row["Remarks"] = f"On Account - invoice {ref} not in current data (old)"
        else:
            row["Remarks"] = "On Account (no invoice)"
        if not co:
            row["Remarks"] += " - SET COMPANY/LOCATION"


# ── Google Sheets ──────────────────────────────────────────────────────────────

def authorize() -> Any:
    creds_path = os.environ["GOOGLE_CREDENTIALS_FILE"]
    token_path = os.environ["GOOGLE_TOKEN_FILE"]
    creds: Credentials | None = None
    if Path(token_path).exists():
        creds = Credentials.from_authorized_user_file(token_path, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(creds_path, SCOPES)
            creds = flow.run_local_server(port=0)
        Path(token_path).write_text(creds.to_json(), encoding="utf-8")
    import httplib2
    from google_auth_httplib2 import AuthorizedHttp
    authed_http = AuthorizedHttp(creds, http=httplib2.Http(timeout=120))
    return build("sheets", "v4", http=authed_http, cache_discovery=False)


def extract_sheet_id(url: str) -> str:
    m = SHEET_ID_RE.search(url)
    if not m:
        raise SystemExit(f"ERROR: could not extract sheet ID from '{url}'")
    return m.group(1)


def fetch_master_names(svc: Any) -> set[str]:
    """Customer names from the credit-limit master (for the unmatched-name check)."""
    url = os.environ.get("CREDIT_LIMIT_SHEET_URL", "").strip()
    if not url:
        return set()
    tab = os.environ.get("CREDIT_LIMIT_SHEET_TAB", "Sheet1")
    try:
        resp = svc.spreadsheets().values().get(
            spreadsheetId=extract_sheet_id(url), range=f"'{tab}'",
            majorDimension="ROWS",
        ).execute()
    except Exception as exc:  # noqa: BLE001 — best-effort check only
        print(f"  (skipped master-name check: {exc})")
        return set()
    values = resp.get("values", [])
    if not values:
        return set()
    header = [h.strip().lower() for h in values[0]]
    name_idx = next((j for j, h in enumerate(header) if "name" in h), None)
    if name_idx is None:
        return set()
    return {
        row[name_idx].strip().upper()
        for row in values[1:]
        if name_idx < len(row) and row[name_idx].strip()
    }


def write_sheet(svc: Any, sheet_id: str, tab: str, rows: list[dict[str, str]]) -> None:
    values = [HEADER] + [[row[h] for h in HEADER] for row in rows]
    svc.spreadsheets().values().clear(spreadsheetId=sheet_id, range=f"'{tab}'").execute()
    svc.spreadsheets().values().update(
        spreadsheetId=sheet_id, range=f"'{tab}'!A1",
        valueInputOption="RAW", body={"values": values},
    ).execute()


def create_sheet(svc: Any, title: str) -> tuple[str, str, str]:
    meta = svc.spreadsheets().create(
        body={"properties": {"title": title}}, fields="spreadsheetId,spreadsheetUrl,sheets.properties.title",
    ).execute()
    tab = meta["sheets"][0]["properties"]["title"]
    return meta["spreadsheetId"], meta["spreadsheetUrl"], tab


def main() -> int:
    parser = argparse.ArgumentParser(description="Import OTHER PAYMENTS xlsx → flat Google Sheet.")
    parser.add_argument("--input", default=str(DEFAULT_XLSX), help="Path to the source .xlsx")
    parser.add_argument("--sheet-url", default=None, help="Existing sheet URL to write into (else a new sheet is created)")
    parser.add_argument("--tab", default="Sheet1", help="Tab name when --sheet-url is given")
    parser.add_argument("--dry-run", action="store_true", help="Parse and print only; write nothing")
    args = parser.parse_args()

    load_dotenv()
    rows = parse_other_payments(Path(args.input))

    svc = authorize()
    maps = build_resolution_maps(svc)
    resolve_ledgers(rows, maps)

    n_against = sum(1 for r in rows if r["Allocation Type"] == "AGAINST_INVOICE")
    n_onacct = sum(1 for r in rows if r["Allocation Type"] == "ON_ACCOUNT")
    n_old = sum(1 for r in rows if r["Allocation Type"] == "ON_ACCOUNT" and r["Ref Invoice No"])
    n_blankco = sum(1 for r in rows if not r["Company"])
    total = sum(float(r["Payment Amount"]) for r in rows if r["Payment Amount"])
    customers = sorted({r["Customer Name"] for r in rows})

    print(f"Parsed {len(rows)} other-payment rows "
          f"({n_against} against-invoice, {n_onacct} on-account; {n_old} old-invoice->on-account) "
          f"across {len(customers)} customers; total Rs {total:,.0f}")

    # Unmatched-name check against the customer master.
    master_names = set(maps["master"].keys())
    if master_names:
        unmatched = sorted(n for n in customers if n.upper() not in master_names)
        if unmatched:
            print(f"  WARNING: {len(unmatched)} customer name(s) not found in master: {unmatched}")
        else:
            print("  All customer names matched the credit-limit master.")
    if n_blankco:
        print(f"  WARNING: {n_blankco} row(s) have no Company resolved (see 'SET COMPANY' remarks).")

    if args.dry_run:
        for r in rows:
            print(f"  {r['Customer Name'][:30]:30} | {r['Company']:10} {r['Location']:7}| "
                  f"{r['Allocation Type']:15} | {r['Ref Invoice No'][:20]:20} | "
                  f"{r['Payment Amount']:>13} | {r['Remarks']}")
        print("\n(dry-run: nothing written)")
        return 0

    if args.sheet_url:
        sheet_id = extract_sheet_id(args.sheet_url)
        tab = args.tab
        url = args.sheet_url
    else:
        sheet_id, url, tab = create_sheet(svc, "Orange Receivables — Other Payments")

    write_sheet(svc, sheet_id, tab, rows)
    print(f"\nWrote {len(rows)} rows to tab '{tab}'.")
    print(f"  Sheet URL: {url}")
    print(f"\nNext: set OTHER_PAYMENTS_SHEET_URL={url}")
    print(f"      and OTHER_PAYMENTS_SHEET_TAB={tab}  in the Orange Receivables Hub .env")
    return 0


if __name__ == "__main__":
    sys.exit(main())
